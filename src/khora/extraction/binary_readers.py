"""Binary format extraction for ingestion.

Converts Excel, Word, and Parquet binary formats to text/markdown so they
can be fed into the standard text-ingest pipeline.  Extractors degrade
gracefully with a warning when optional dependencies are missing.

PDF extraction is not bundled with khora.  Callers passing a ``.pdf`` path
to :func:`extract_if_needed` get a :class:`NotImplementedError` — preprocess
PDFs upstream (extract text to ``.txt``/``.md``) or use ``khora-cli``'s PDF
preprocessing, then pass the extracted text to ``remember()``.

Install: ``pip install khora[binary-readers]`` for openpyxl + python-docx,
and ``pip install khora[parquet]`` for pyarrow.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from loguru import logger

# ---------------------------------------------------------------------------
# Optional dependency detection
# ---------------------------------------------------------------------------

_HAS_OPENPYXL = False
try:
    import openpyxl  # noqa: F401

    _HAS_OPENPYXL = True
except ImportError:
    pass

_HAS_DOCX = False
try:
    import docx  # noqa: F401

    _HAS_DOCX = True
except ImportError:
    pass

_HAS_PYARROW = False
try:
    import pyarrow  # noqa: F401

    _HAS_PYARROW = True
except ImportError:
    pass


# ---------------------------------------------------------------------------
# PDF sentinel
# ---------------------------------------------------------------------------

_PDF_REMOVED_MESSAGE = (
    "khora does not include PDF parsing. "
    "Preprocess PDFs upstream (e.g. extract text to a .txt or .md file) "
    "or use khora-cli's PDF preprocessing, then pass the extracted text to remember()."
)


def _pdf_not_supported(path: Path) -> str:
    """Dispatcher entry that raises for ``.pdf`` paths.

    PDF parsing is not bundled with khora; see the module docstring.
    """
    raise NotImplementedError(_PDF_REMOVED_MESSAGE)


# ---------------------------------------------------------------------------
# Extractors
# ---------------------------------------------------------------------------


def extract_xlsx_text(path: Path) -> str:
    """Extract text from an Excel (.xlsx) file as CSV-formatted text.

    Uses openpyxl for reading. Returns empty string if not installed.
    """
    if not _HAS_OPENPYXL:
        logger.warning(f"openpyxl not installed — cannot extract text from {path.name}. Install: pip install openpyxl")
        return ""

    import openpyxl

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sections: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            buf = io.StringIO()
            writer = csv.writer(buf)
            row_count = 0

            for row in ws.iter_rows(values_only=True):
                writer.writerow([str(cell) if cell is not None else "" for cell in row])
                row_count += 1

            if row_count > 0:
                sections.append(f"## Sheet: {sheet_name}\n\n```csv\n{buf.getvalue()}```\n")

        wb.close()
        return "\n".join(sections)
    except Exception as e:
        logger.error(f"Excel extraction failed for {path.name}: {e}")
        return ""


def extract_xls_text(path: Path) -> str:
    """Extract text from a legacy .xls file.

    Falls back to a warning — xlrd is not included as a dependency.
    """
    logger.warning(f"Legacy .xls format not supported: {path.name}. Convert to .xlsx first.")
    return ""


def extract_docx_text(path: Path) -> str:
    """Extract text from a Word (.docx) file.

    Uses python-docx for paragraph extraction.
    Returns empty string if python-docx is not installed.
    """
    if not _HAS_DOCX:
        logger.warning(
            f"python-docx not installed — cannot extract text from {path.name}. Install: pip install python-docx"
        )
        return ""

    import docx

    try:
        doc = docx.Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except Exception as e:
        logger.error(f"DOCX extraction failed for {path.name}: {e}")
        return ""


def extract_parquet_text(path: Path) -> str:
    """Extract data from a Parquet file as markdown table.

    Uses pyarrow for reading. Returns the first 100 rows and schema.
    Returns empty string if pyarrow is not installed.
    """
    if not _HAS_PYARROW:
        logger.warning(f"pyarrow not installed — cannot extract data from {path.name}. Install: pip install pyarrow")
        return ""

    import pyarrow.parquet as pq

    try:
        table = pq.read_table(str(path))
        schema_text = "## Schema\n\n"
        for field in table.schema:
            schema_text += f"- **{field.name}**: {field.type}\n"

        # Convert first 100 rows to pandas for display
        df = table.to_pandas().head(100)
        rows_text = f"\n## Data ({len(table)} total rows, showing first {min(100, len(table))})\n\n"
        rows_text += df.to_markdown(index=False) if hasattr(df, "to_markdown") else df.to_string(index=False)

        return schema_text + rows_text
    except Exception as e:
        logger.error(f"Parquet extraction failed for {path.name}: {e}")
        return ""


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

#: Map of file extensions to extraction functions
_EXTRACTORS: dict[str, callable] = {
    ".pdf": _pdf_not_supported,  # raises NotImplementedError — not a text extractor
    ".xlsx": extract_xlsx_text,
    ".xls": extract_xls_text,
    ".docx": extract_docx_text,
    ".parquet": extract_parquet_text,
}


def get_extraction_warning(path: Path) -> str | None:
    """Return a user-facing warning if extraction would fail for this file type.

    Checks whether the required optional dependency is installed for the
    given file extension.  Returns ``None`` when everything looks fine.
    """
    ext = path.suffix.lower()
    if ext == ".pdf":
        return f"Cannot extract text from {path.name} — {_PDF_REMOVED_MESSAGE}"
    if ext in (".xlsx", ".xls") and not _HAS_OPENPYXL:
        return f"Cannot extract text from {path.name} — install openpyxl: pip install openpyxl"
    if ext == ".docx" and not _HAS_DOCX:
        return f"Cannot extract text from {path.name} — install python-docx: pip install python-docx"
    if ext == ".parquet" and not _HAS_PYARROW:
        return f"Cannot extract text from {path.name} — install pyarrow: pip install pyarrow"
    return None


def extract_if_needed(path: Path) -> Path | None:
    """Extract text from a binary file if an extractor is available.

    If extraction succeeds, writes a `{name}_extracted.md` file alongside
    the original and returns the path to it.  Returns None if no
    extraction was needed or if it failed.

    Raises :class:`NotImplementedError` for ``.pdf`` inputs: PDF parsing is
    not bundled with khora; preprocess PDFs upstream.
    """
    ext = path.suffix.lower()
    extractor = _EXTRACTORS.get(ext)

    if extractor is None:
        return None

    text = extractor(path)
    if not text:
        return None

    # Write extracted text alongside the original
    extracted_path = path.with_name(f"{path.stem}_extracted.md")
    extracted_path.write_text(text, encoding="utf-8")
    logger.info(f"Extracted {len(text):,} chars from {path.name} → {extracted_path.name}")
    return extracted_path
